
# coding: utf-8

# In[1]:


import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import sys
import os

import xlsxwriter
from win32com.client import Dispatch

from tkinter import *
from tkinter import ttk
from tkinter import filedialog


# In[2]:


def set_border(ws, cell_range):
#     rows = ws.range(cell_range)
    thin_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    
    for rows in ws[cell_range]:
        for cell in rows:
            cell.border = thin_border
#             row[0].style.borders.left.border_style = Border.
#             row[-1].style.borders.right.border_style = Border.BORDER_MEDIUM
#         for c in rows[0]:
#             c.style.borders.top.border_style = Border.BORDER_MEDIUM
#         for c in rows[-1]:
#             c.style.borders.bottom.border_style = Border.BORDER_MEDIUM


# In[3]:


def colorCell(row, column, value):
    cell = sheet.cell(row=row, column=column)
    if(value != None):
        cell.value = value#depotHeaderName+" Depot Route List"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
    
    thin_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    
    cell.border = thin_border


# In[4]:


root = Tk()
root.filename = filedialog.askopenfilename(filetypes = (("Excel files","*.xlsx"),("All files", "*.*")))
user_input_path = root.filename

root.destroy()

df_main = pd.read_excel(user_input_path)

if("Depot" not in df_main):
    print("No column with name 'Depot' was found in " + excel)
    print("Add/Modify a column with name 'Depot' and re-run Depotwise-MasterToSrcDes_With StopCode,LatLong-v4.0")
    pWadalrint()
    if(input("Press enter to exit..")):
        sys.exit()
else:
    depotList = df_main["Depot"]
    depotList = depotList.tolist()
    #depotList = (map(str.capitalize, depotList.tolist()).tolist()
    
    condition = True
    while(condition):
        print()
        user_input_depot = input("Which depot do you want to process?(case sensitive): ")

        depotName = user_input_depot.capitalize()
        # print(depotHeaderName)

        if(depotName in depotList):
            condition = False
            df_main_depot_specific = df_main.loc[ df_main["Depot"] == depotName ].copy()
        else:
            print("Entered depot - \""+depotName+"\" is not found in excel. Try again..")

print()
print("Processing.. please wait..")
# In[5]:



# depotHeaderName = user_input_path.split('\\')[-1].split('.')[0].split(' ')[0]
depotHeaderName = user_input_depot
# print(depotHeaderName)


# In[6]:


#Store the output file path in a variable
output_file = os.getcwd()+"\\"+depotHeaderName+" output.xlsx"


# In[7]:


df_depot_specific_full = df_main.loc[(df_main['Depot'] == user_input_depot)]


# In[8]:


#Retain only the required columns in the dataframe
# df_main_copy = df_main.copy()
df_depot_specific = df_depot_specific_full[['Route Name', 'Stop Serial', 'Stop Code', 'Stop Name Lang 2', 'Stop Name Lang 1', 'Latitude', 'Longitude']].copy()


# In[9]:



#Extract the Route Names to a List of unique values

routeList = df_depot_specific['Route Name'].unique().tolist()


# In[10]:


#Extract data for first and last stops into two dataframes respectivly.

writeRow = 1
result = pd.DataFrame()
df_FirstStop = pd.DataFrame()
df_LastStop = pd.DataFrame()
for x in routeList:
    df_FirstStop = df_FirstStop.append(df_depot_specific.loc[ (df_depot_specific['Stop Serial'] == 1) & (df_depot_specific['Route Name'] == x) ], sort=False)

    count = int((df_depot_specific.loc[(df_depot_specific['Route Name'] == x)]).groupby('Route Name').size())
    df_LastStop = df_LastStop.append(df_depot_specific.loc[ (df_depot_specific['Stop Serial'] == count) & (df_depot_specific['Route Name'] == x) ], sort=False)
       
# del df_FirstStop['Stop Serial']
# del df_LastStop['Stop Serial']
# print(tdff2)
# result.merge(tdff1,tdff2,on='Route Name',ignore_index=True)


# In[11]:


#Merge the two dataframes into one dataframe

result = pd.merge(df_FirstStop,df_LastStop, on = "Route Name")


# In[13]:


#Write dataframe to excel

result.to_excel(output_file,index=False,header=False,startcol=0,startrow=3, sheet_name="Source and Destination")

# df_depot_specific_full.to_excel(output_file,index=False,header=True,startcol=0,startrow=0, sheet_name="Org Data")

# In[15]:



#Adding the headers here


# To open the workbook  
# workbook object is created 
book = load_workbook(output_file)

# Get workbook active sheet object 
# from the active attribute 
sheet = book.active

# sheet.merge_cells('A1:I1')
sheet.merge_cells(start_row=1,start_column=1,end_row=1,end_column=13)   #... Depot Route List
sheet.merge_cells(start_row=2,start_column=1,end_row=3,end_column=1)   #Route Name
sheet.merge_cells(start_row=2,start_column=2,end_row=2,end_column=7)   #From
sheet.merge_cells(start_row=2,start_column=8,end_row=2,end_column=13)   #To


colorCell(1, 1, depotHeaderName+" Depot Route List")
colorCell(2, 1, 'Route Name')
colorCell(2, 2, 'From')
colorCell(2, 8, 'To')
colorCell(3, 2, 'Stop Serial')
colorCell(3, 3, 'Stop Code')
colorCell(3, 4, 'Stop Name Lang 2')
colorCell(3, 5, 'Stop Name Lang 1')
colorCell(3, 6, 'Latitude')
colorCell(3, 7, 'Longitude')
colorCell(3, 8, 'Stop Serial')
colorCell(3, 9, 'Stop Code')
colorCell(3, 10, 'Stop Name Lang 2')
colorCell(3, 11, 'Stop Name Lang 1')
colorCell(3, 12, 'Latitude')
colorCell(3, 13, 'Longitude')

set_border(sheet,"A1:M3")
# cell = sheet.cell(row=3,column=5)
# # cell.border = thin_border
# cell .value = 'Marathi'
# cell.alignment = Alignment(horizontal='center', vertical='center')
# cell.fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')

#Create new sheet and write the complete data for the requested depot
book.save(output_file)


# In[16]:

orgDataSheetName = str(depotHeaderName) + " - Route Data"
book.create_sheet(orgDataSheetName)
sheet = book[orgDataSheetName]

#Take only first 8 coloumns to the "<depot name> - Route Data" sheet
for r in dataframe_to_rows(df_depot_specific_full.iloc[:,[0,1,2,3,4,5,6,7]], index=False, header=True):
    sheet.append(r)

i=1
for x in sheet.columns:
    colorCell(1, i, None)
    i=i+1
book.save(output_file)


# In[17]:


#Make all columns in the output excel to AutoFit to the data.

excel = Dispatch('Excel.Application')
wb = excel.Workbooks.Open(output_file)

#Activate second sheet
excel.Worksheets(2).Activate()

#Autofit column in active sheet
excel.ActiveSheet.Columns.AutoFit()

#Save changes in a new file
#wb.SaveAs("D:\\output_fit.xlsx")

#Or simply save changes in a current file
wb.Save()

wb.Close()


#Make all columns in the output excel to AutoFit to the data.

excel = Dispatch('Excel.Application')
wb = excel.Workbooks.Open(output_file)

#Activate first sheet
excel.Worksheets(1).Activate()

#Autofit column in active sheet
excel.ActiveSheet.Columns.AutoFit()

#Save changes in a new file
#wb.SaveAs("D:\\output_fit.xlsx")

#Or simply save changes in a current file
wb.Save()

wb.Close()



if(input("Completed processing. Press enter to exit..")):
        sys.exit()
