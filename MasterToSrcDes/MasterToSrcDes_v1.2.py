
# coding: utf-8

# In[1]:


import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import sys
import os

import xlsxwriter
from win32com.client import Dispatch


# In[2]:


def set_border(ws, cell_range):
#     rows = ws.range(cell_range)
    thin_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    
    for rows in ws[cell_range]:
        for cell in rows:
            cell.border = thin_border
#             row[0].style.borders.left.border_style = Border.
#             row[-1].style.borders.right.border_style = Border.BORDER_MEDIUM
#         for c in rows[0]:
#             c.style.borders.top.border_style = Border.BORDER_MEDIUM
#         for c in rows[-1]:
#             c.style.borders.bottom.border_style = Border.BORDER_MEDIUM


# In[3]:


def colorCell(row, column, value):
    cell = sheet.cell(row=row, column=column)
    if(value != None):
        cell.value = value#depotHeaderName+" Depot Route List"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
    
    thin_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    
    cell.border = thin_border


# In[4]:


user_input_path = input("Enter the path of excel file: ")
# assert os.path.exists(user_input), "I did not find the file at, "+str(user_input)

user_input_depot = input("Which depot do you want to process?(case sensitive): ")

# depotHeaderName = user_input_path.split('\\')[-1].split('.')[0].split(' ')[0]
depotHeaderName = user_input_depot
# print(depotHeaderName)


# In[5]:


#Store the output file path in a variable
output_file = os.getcwd()+"\\"+depotHeaderName+" output.xlsx"


# In[6]:


#Read the excel data into dataframe

df_main = pd.read_excel(user_input_path)


# In[7]:


df_depot_specific_full = df_main.loc[(df_main['Depot'] == user_input_depot)]


# In[8]:


#Retain only the required columns in the dataframe
# df_main_copy = df_main.copy()
df_depot_specific = df_depot_specific_full[['Route Name', 'Stop Serial', 'English Stop Name', 'Marathi Stop Name']].copy()


# In[9]:


#Extract the Route Names to a List of unique values

routeList = df_depot_specific['Route Name'].unique().tolist()


# In[10]:


#Extract data for first and last stops into two dataframes respectivly.

writeRow = 1
result = pd.DataFrame()
df_FirstStop = pd.DataFrame()
df_LastStop = pd.DataFrame()
for x in routeList:
    df_FirstStop = df_FirstStop.append(df_depot_specific.loc[ (df_depot_specific['Stop Serial'] == 1) & (df_depot_specific['Route Name'] == x) ], sort=False)

    count = int((df_depot_specific.loc[(df_depot_specific['Route Name'] == x)]).groupby('Route Name').size())
    df_LastStop = df_LastStop.append(df_depot_specific.loc[ (df_depot_specific['Stop Serial'] == count) & (df_depot_specific['Route Name'] == x) ], sort=False)
    
    
del df_FirstStop['Stop Serial']
del df_LastStop['Stop Serial']
# print(tdff2)
# result.merge(tdff1,tdff2,on='Route Name',ignore_index=True)


# In[11]:


#Merge the two dataframes into one dataframe

result = pd.merge(df_FirstStop,df_LastStop, on = "Route Name")


# In[12]:


#Write dataframe to excel

result.to_excel(output_file,index=False,header=False,startcol=0,startrow=3, sheet_name="FromToData")

# df_depot_specific_full.to_excel(output_file,index=False,header=True,startcol=0,startrow=0, sheet_name="Org Data")


# In[13]:


#Make all columns in the output excel to AutoFit to the data.

excel = Dispatch('Excel.Application')
wb = excel.Workbooks.Open(output_file)

#Activate first sheet
excel.Worksheets(1).Activate()

#Autofit column in active sheet
excel.ActiveSheet.Columns.AutoFit()

#Save changes in a new file
#wb.SaveAs("D:\\output_fit.xlsx")

#Or simply save changes in a current file
wb.Save()

wb.Close()


# In[14]:


#Adding the headers here


# To open the workbook  
# workbook object is created 
book = load_workbook(output_file)

# Get workbook active sheet object 
# from the active attribute 
sheet = book.active

# sheet.merge_cells('A1:E1')
sheet.merge_cells(start_row=1,start_column=1,end_row=1,end_column=5)   #... Depot Route List
sheet.merge_cells(start_row=2,start_column=1,end_row=3,end_column=1)   #Route Name
sheet.merge_cells(start_row=2,start_column=2,end_row=2,end_column=3)   #From
sheet.merge_cells(start_row=2,start_column=4,end_row=2,end_column=5)   #To


colorCell(1, 1, depotHeaderName+" Depot Route List")
colorCell(2, 1, 'Route Name')
colorCell(2, 2, 'From')
colorCell(2, 4, 'To')
colorCell(3, 2, 'English')
colorCell(3, 3, 'Marathi')
colorCell(3, 4, 'English')
colorCell(3, 5, 'Marathi')

set_border(sheet,"A1:E3")
# cell = sheet.cell(row=3,column=5)
# # cell.border = thin_border
# cell .value = 'Marathi'
# cell.alignment = Alignment(horizontal='center', vertical='center')
# cell.fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')

#Create new sheet and write the complete data for the requested depot
book.save(output_file)


# In[15]:


book.create_sheet("Org Data")
sheet = book["Org Data"]

for r in dataframe_to_rows(df_depot_specific_full, index=True, header=True):
    sheet.append(r)

i=1
for x in sheet.columns:
    colorCell(1, i, None)
    i=i+1
book.save(output_file)


# In[16]:


#Make all columns in the output excel to AutoFit to the data.

excel = Dispatch('Excel.Application')
wb = excel.Workbooks.Open(output_file)

#Activate second sheet
excel.Worksheets(2).Activate()

#Autofit column in active sheet
excel.ActiveSheet.Columns.AutoFit()

#Save changes in a new file
#wb.SaveAs("D:\\output_fit.xlsx")

#Or simply save changes in a current file
wb.Save()

wb.Close()

